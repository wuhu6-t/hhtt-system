import os
import shutil
import zipfile
import tempfile
import uuid
import time
from flask import Blueprint, render_template, request, jsonify, send_file
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.workbook.views import BookView
from openpyxl.worksheet.views import Selection
import xml.etree.ElementTree as ET

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
fujian_app = Blueprint('fujian', __name__, template_folder=os.path.join(BASE_DIR, 'templates', 'fujian'))

fujian_app.config = {
    'OUTPUT_BASE': os.path.join(os.path.dirname(os.path.abspath(__file__)), 'hhtt-fujian', '提交')
}

ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.png', '.jpg', '.jpeg', '.docx', '.doc', '.pdf', '.txt', '.zip', '.rar'}

district_prefix_map = {
    '个旧': '个旧市', '元阳': '元阳县', '屏边': '屏边县',
    '建水': '建水县', '开远': '开远市', '弥勒': '弥勒市',
    '河口': '河口县', '泸西': '泸西县', '红河': '红河县',
    '绿春': '绿春县', '金平': '金平县',
}

task_results = {}


def generate_task_id():
    timestamp = int(time.time() * 1000)
    uuid_str = str(uuid.uuid4())[:8]
    return f'{timestamp}-{uuid_str}'


def fix_view_settings(xlsx_path):
    try:
        temp_path = xlsx_path + '.tmp'
        with zipfile.ZipFile(xlsx_path, 'r') as z_in:
            with zipfile.ZipFile(temp_path, 'w') as z_out:
                for item in z_in.infolist():
                    if item.filename == 'xl/worksheets/sheet1.xml':
                        xml_content = z_in.read(item.filename).decode('utf-8')
                        root = ET.fromstring(xml_content)
                        ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        sheet_view = root.find('.//sheetView', ns)
                        if sheet_view is not None:
                            sheet_view.set('topLeftCell', 'A1')
                            frozen_pane = sheet_view.find('.//pane', ns)
                            if frozen_pane is not None:
                                sheet_view.remove(frozen_pane)
                            for selection in sheet_view.findall('.//selection', ns):
                                selection.set('activeCell', 'A1')
                                selection.set('sqref', 'A1')
                        else:
                            worksheet = root.find('.//worksheet', ns)
                            if worksheet is not None:
                                sheet_view = ET.SubElement(worksheet, 'sheetView')
                                sheet_view.set('topLeftCell', 'A1')
                                selection = ET.SubElement(sheet_view, 'selection')
                                selection.set('activeCell', 'A1')
                                selection.set('sqref', 'A1')
                        xml_content = ET.tostring(root, encoding='utf-8', method='xml').decode('utf-8')
                        z_out.writestr(item.filename, xml_content)
                    else:
                        z_out.writestr(item, z_in.read(item.filename))
        os.remove(xlsx_path)
        os.rename(temp_path, xlsx_path)
        return True
    except Exception:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        return False


def split_by_district(input_file, output_base):
    results = []
    try:
        new_wb = load_workbook(input_file)
        new_ws = new_wb.active

        total_rows = new_ws.max_row
        districts = []
        for row in range(2, total_rows + 1):
            district = new_ws.cell(row=row, column=2).value
            if district and district not in districts:
                districts.append(district)

        districts = sorted(districts)

        for district in districts:
            try:
                new_wb = load_workbook(input_file)
                new_ws = new_wb.active

                for row in range(new_ws.max_row, 1, -1):
                    cell_value = new_ws.cell(row=row, column=2).value
                    if cell_value != district:
                        new_ws.delete_rows(row)

                last_data_row = new_ws.max_row

                merged_to_remove = []
                for merge in list(new_ws.merged_cells.ranges):
                    try:
                        if merge.min_row > last_data_row:
                            merged_to_remove.append(merge)
                    except:
                        pass
                for merge in merged_to_remove:
                    try:
                        new_ws.unmerge_cells(str(merge))
                    except:
                        pass

                rows_to_remove = []
                for row_idx in list(new_ws.row_dimensions.keys()):
                    if row_idx > last_data_row:
                        rows_to_remove.append(row_idx)
                for row_idx in rows_to_remove:
                    try:
                        del new_ws.row_dimensions[row_idx]
                    except:
                        pass

                start_row = last_data_row + 4

                new_ws.merge_cells(f'B{start_row}:D{start_row}')
                cell = new_ws.cell(row=start_row, column=2, value='经办人：')
                cell.alignment = Alignment(horizontal='right')

                new_ws.merge_cells(f'H{start_row}:J{start_row}')
                cell = new_ws.cell(row=start_row, column=8, value='经办人：')
                cell.alignment = Alignment(horizontal='right')

                new_ws.merge_cells(f'B{start_row + 2}:D{start_row + 2}')
                cell = new_ws.cell(row=start_row + 2, column=2, value='区县分管领导意见及盖章：')
                cell.alignment = Alignment(horizontal='right')

                new_ws.merge_cells(f'H{start_row + 2}:J{start_row + 2}')
                cell = new_ws.cell(row=start_row + 2, column=8, value='区县分管领导意见及盖章：')
                cell.alignment = Alignment(horizontal='right')

                new_ws.merge_cells(f'B{start_row + 4}:D{start_row + 4}')
                cell = new_ws.cell(row=start_row + 4, column=2, value='日期：')
                cell.alignment = Alignment(horizontal='right')

                new_ws.merge_cells(f'H{start_row + 4}:J{start_row + 4}')
                cell = new_ws.cell(row=start_row + 4, column=8, value='日期：')
                cell.alignment = Alignment(horizontal='right')

                new_ws.sheet_view.topLeftCell = 'A1'
                new_ws.sheet_view.selection = [Selection(activeCell='A1', sqref='A1')]
                new_wb.views = [BookView(activeTab=0)]

                prefix = district[:2]
                folder_path = os.path.join(output_base, f'1{prefix}-提交附件', '分摊比例表')
                os.makedirs(folder_path, exist_ok=True)
                output_file = os.path.join(folder_path, '分摊比例表.xlsx')

                new_wb.save(output_file)
                fix_view_settings(output_file)

                results.append({'district': district, 'success': True, 'file': output_file, 'rows': new_ws.max_row - 1})
            except Exception as e:
                results.append({'district': district, 'success': False, 'error': str(e)})

        return results
    except Exception as e:
        return [{'district': '整体', 'success': False, 'error': str(e)}]


def classify_and_copy_files(source_folder, output_base):
    results = []
    exclude_files = ['分摊比例表.xlsx', '系统无订单情况说明.xlsx']
    
    all_files = []
    for root, dirs, filenames in os.walk(source_folder):
        for filename in filenames:
            if filename in exclude_files:
                continue
            file_path = os.path.join(root, filename)
            rel_path = os.path.relpath(file_path, source_folder)
            all_files.append({'name': filename, 'path': rel_path, 'full_path': file_path})

    for file_info in all_files:
        filename = file_info['name']
        prefix = filename[:2]
        
        if prefix in district_prefix_map:
            district = district_prefix_map[prefix]
            target_folder = os.path.join(output_base, f'1{prefix}-提交附件', '分摊比例表')
            os.makedirs(target_folder, exist_ok=True)
            
            source_path = file_info['full_path']
            target_path = os.path.join(target_folder, filename)
            
            try:
                shutil.copy2(source_path, target_path)
                results.append({'file': filename, 'district': district, 'success': True})
            except Exception as e:
                results.append({'file': filename, 'district': district, 'success': False, 'error': str(e)})
        else:
            target_folder = os.path.join(output_base, '其他文件')
            os.makedirs(target_folder, exist_ok=True)
            
            source_path = file_info['full_path']
            target_path = os.path.join(target_folder, filename)
            
            try:
                shutil.copy2(source_path, target_path)
                results.append({'file': filename, 'district': '其他', 'success': True})
            except Exception as e:
                results.append({'file': filename, 'district': '其他', 'success': False, 'error': str(e)})

    return results


@fujian_app.route('/')
def index():
    template_path = os.path.join(BASE_DIR, 'templates', 'fujian', 'index.html')
    with open(template_path, 'r', encoding='utf-8') as f:
        return f.read()


@fujian_app.route('/upload', methods=['POST'])
def upload_files():
    task_id = generate_task_id()
    
    try:
        temp_dir = tempfile.mkdtemp()
        task_output_dir = os.path.join(fujian_app.config['OUTPUT_BASE'], task_id)
        os.makedirs(task_output_dir, exist_ok=True)
        
        if 'files' not in request.files:
            return jsonify({'success': False, 'message': '请选择文件', 'task_id': task_id})
        
        files = request.files.getlist('files')
        uploaded_file_names = []
        has_excel = False
        
        for file in files:
            if file.filename:
                relative_path = getattr(file, 'relative_path', file.filename)
                file_path = os.path.join(temp_dir, relative_path)
                os.makedirs(os.path.dirname(file_path), exist_ok=True)
                file.save(file_path)
                uploaded_file_names.append({'name': file.filename, 'path': relative_path})
                
                if file.filename == '分摊比例表.xlsx':
                    has_excel = True
        
        if not uploaded_file_names:
            return jsonify({'success': False, 'message': '未上传任何文件', 'task_id': task_id})
        
        results = {'excel': [], 'files': []}
        
        excel_path = os.path.join(temp_dir, '分摊比例表.xlsx')
        if has_excel and os.path.exists(excel_path):
            excel_results = split_by_district(excel_path, task_output_dir)
            results['excel'] = excel_results
        elif not has_excel:
            for root, dirs, filenames in os.walk(temp_dir):
                for filename in filenames:
                    if filename == '分摊比例表.xlsx':
                        excel_path = os.path.join(root, filename)
                        excel_results = split_by_district(excel_path, task_output_dir)
                        results['excel'] = excel_results
                        has_excel = True
                        break
                if has_excel:
                    break
        
        file_results = classify_and_copy_files(temp_dir, task_output_dir)
        results['files'] = file_results
        
        success_count = sum(1 for r in results['excel'] if r.get('success')) + sum(1 for r in results['files'] if r.get('success'))
        total_count = len(results['excel']) + len(results['files'])
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        task_results[task_id] = {
            'output_dir': task_output_dir,
            'success': True,
            'timestamp': time.time()
        }
        
        message = f'处理完成！成功 {success_count}/{total_count}'
        return jsonify({'success': True, 'message': message, 'results': results, 'task_id': task_id})
    
    except Exception as e:
        task_results[task_id] = {
            'output_dir': None,
            'success': False,
            'timestamp': time.time()
        }
        return jsonify({'success': False, 'message': f'处理失败：{str(e)}', 'task_id': task_id})


@fujian_app.route('/download/<task_id>', methods=['GET'])
def download_files(task_id):
    try:
        if task_id not in task_results:
            return jsonify({'success': False, 'message': '任务不存在或已过期'}), 404
        
        task_info = task_results[task_id]
        output_dir = task_info['output_dir']
        
        if not output_dir or not os.path.exists(output_dir):
            return jsonify({'success': False, 'message': '没有可下载的文件'}), 404
        
        temp_dir = tempfile.mkdtemp()
        zip_filename = f'提交_{task_id}.zip'
        zip_path = os.path.join(temp_dir, zip_filename)
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, dirs, filenames in os.walk(output_dir):
                for filename in filenames:
                    file_path = os.path.join(root, filename)
                    rel_path = os.path.relpath(file_path, output_dir)
                    zf.write(file_path, rel_path)
        
        return send_file(zip_path, as_attachment=True, attachment_filename='提交.zip')
    
    except Exception as e:
        return jsonify({'success': False, 'message': f'下载失败：{str(e)}'}), 500


@fujian_app.route('/cleanup', methods=['POST'])
def cleanup_old_tasks():
    try:
        cutoff_time = time.time() - 3600
        cleaned_count = 0
        
        for task_id, task_info in list(task_results.items()):
            if task_info['timestamp'] < cutoff_time:
                output_dir = task_info.get('output_dir')
                if output_dir and os.path.exists(output_dir):
                    shutil.rmtree(output_dir, ignore_errors=True)
                del task_results[task_id]
                cleaned_count += 1
        
        return jsonify({'success': True, 'cleaned': cleaned_count})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})