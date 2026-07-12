import os
import tempfile
import shutil
import uuid
from flask import Blueprint, render_template, request, jsonify, send_from_directory
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import datetime

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
manage_app = Blueprint('manage', __name__, template_folder=os.path.join(BASE_DIR, 'templates', 'manage'), static_folder=os.path.join(BASE_DIR, 'static', 'manage'))
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'hhtt-manage', 'uploads')
OUTPUT_FOLDER = os.path.join(BASE_DIR, 'hhtt-manage', 'outputs')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


def get_column_index(ws, column_name):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == column_name:
            return col
    return None


def get_payment_codes(ws, column_index):
    codes = set()
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=column_index).value
        if value:
            codes.add(str(value).strip())
    return codes


def filter_rows_by_codes(ws, column_index, valid_codes):
    rows_to_delete = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=column_index).value
        code = str(value).strip() if value else ""
        if code not in valid_codes:
            rows_to_delete.append(row)
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
    return len(rows_to_delete)


def create_allocation_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "分摊比例表"
    headers = [
        "地市", "区县", "站址编码", "铁塔站点名称",
        "报账点编码", "报账点名称", "电表编码",
        "上期分摊比", "上期时间：XX年XX月XX日-XX年XX月XX日",
        "本期分摊比", "本期分摊开始时间：XX年XX月XX日",
        "分摊比例变化原因"
    ]
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="宋体", size=11, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    column_widths = [10, 10, 15, 20, 15, 20, 15, 12, 30, 12, 25, 25]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 40
    return wb, ws


def to_percent(value):
    if value is None or value == "":
        return ""
    try:
        num = float(value)
        return f"{num:.2f}%"
    except (ValueError, TypeError):
        return value


def build_mappings(source_ws):
    county_mapping, station_mapping = {}, {}
    code_col = get_column_index(source_ws, "报账点编码")
    county_col = get_column_index(source_ws, "所属区县")
    station_col = get_column_index(source_ws, "铁塔站址编码")
    if code_col is None:
        return county_mapping, station_mapping
    for row in range(2, source_ws.max_row + 1):
        code = source_ws.cell(row=row, column=code_col).value
        if code:
            code_str = str(code).strip()
            if county_col:
                county_mapping[code_str] = source_ws.cell(row=row, column=county_col).value
            if station_col:
                station_mapping[code_str] = source_ws.cell(row=row, column=station_col).value
    return county_mapping, station_mapping


def vlookup_county_and_fill_city(source_ws, target_ws, county_mapping):
    target_code_col = get_column_index(target_ws, "报账点编码")
    target_county_col = get_column_index(target_ws, "区县")
    target_city_col = get_column_index(target_ws, "地市")
    if not all([target_code_col, target_county_col, target_city_col]):
        return 0
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    filled = 0
    for row in range(2, target_ws.max_row + 1):
        code = target_ws.cell(row=row, column=target_code_col).value
        if code:
            code_str = str(code).strip()
            if code_str in county_mapping and county_mapping[code_str]:
                target_ws.cell(row=row, column=target_county_col, value=county_mapping[code_str]).border = thin_border
                target_ws.cell(row=row, column=target_county_col).alignment = Alignment(horizontal="center", vertical="center")
                target_ws.cell(row=row, column=target_city_col, value="红河").border = thin_border
                target_ws.cell(row=row, column=target_city_col).alignment = Alignment(horizontal="center", vertical="center")
                filled += 1
    return filled


def vlookup_station_code(target_ws, station_mapping):
    target_code_col = get_column_index(target_ws, "报账点编码")
    target_station_col = get_column_index(target_ws, "站址编码")
    if not all([target_code_col, target_station_col]):
        return 0
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    filled = 0
    for row in range(2, target_ws.max_row + 1):
        code = target_ws.cell(row=row, column=target_code_col).value
        if code:
            code_str = str(code).strip()
            if code_str in station_mapping and station_mapping[code_str]:
                target_ws.cell(row=row, column=target_station_col, value=station_mapping[code_str]).border = thin_border
                target_ws.cell(row=row, column=target_station_col).alignment = Alignment(horizontal="center", vertical="center")
                filled += 1
    return filled


def build_station_name_mapping(source_wb):
    mapping = {}
    target_sheets = ["铁塔类", "微站"]
    for sheet_name in target_sheets:
        if sheet_name in source_wb.sheetnames:
            ws = source_wb[sheet_name]
            station_code_col = get_column_index(ws, "站址编码")
            station_name_col = get_column_index(ws, "铁塔站址名称")
            if not all([station_code_col, station_name_col]):
                continue
            for row in range(2, ws.max_row + 1):
                code = ws.cell(row=row, column=station_code_col).value
                name = ws.cell(row=row, column=station_name_col).value
                if code:
                    mapping[str(code).strip()] = name
    return mapping


def vlookup_station_name(target_ws, station_name_mapping):
    target_station_code_col = get_column_index(target_ws, "站址编码")
    target_station_name_col = get_column_index(target_ws, "铁塔站点名称")
    if not all([target_station_code_col, target_station_name_col]):
        return 0
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    filled = 0
    for row in range(2, target_ws.max_row + 1):
        code = target_ws.cell(row=row, column=target_station_code_col).value
        if code:
            code_str = str(code).strip()
            if code_str in station_name_mapping and station_name_mapping[code_str]:
                target_ws.cell(row=row, column=target_station_name_col, value=station_name_mapping[code_str]).border = thin_border
                target_ws.cell(row=row, column=target_station_name_col).alignment = Alignment(horizontal="center", vertical="center")
                filled += 1
    return filled


def parse_date_for_sort(value):
    if value is None or value == "":
        return None
    if hasattr(value, 'year'):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
            try:
                return datetime.datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def format_date(value):
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        value = value.strip()
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
            try:
                date_obj = datetime.datetime.strptime(value, fmt)
                return f"{date_obj.year}年{date_obj.month}月{date_obj.day}日"
            except ValueError:
                continue
        return value
    if hasattr(value, 'year'):
        return f"{value.year}年{value.month}月{value.day}日"
    return str(value)


def build_payment_period_mapping(source_ws):
    mapping = {}
    code_col = get_column_index(source_ws, "报账点编码")
    start_col = get_column_index(source_ws, "缴费期始")
    end_col = get_column_index(source_ws, "缴费期终")
    if not all([code_col, start_col, end_col]):
        return mapping
    records = []
    for row in range(2, source_ws.max_row + 1):
        code = source_ws.cell(row=row, column=code_col).value
        start = source_ws.cell(row=row, column=start_col).value
        end = source_ws.cell(row=row, column=end_col).value
        if code:
            records.append({
                "code": str(code).strip(),
                "start": start,
                "end": end,
                "start_parsed": parse_date_for_sort(start),
            })
    records.sort(key=lambda x: x["start_parsed"] if x["start_parsed"] else datetime.datetime.min, reverse=True)
    seen = set()
    for rec in records:
        code = rec["code"]
        if code not in seen:
            seen.add(code)
            start_str = format_date(rec["start"])
            end_str = format_date(rec["end"])
            if start_str or end_str:
                mapping[code] = f"{start_str}-{end_str}"
    return mapping


def vlookup_payment_period(target_ws, period_mapping):
    target_code_col = get_column_index(target_ws, "报账点编码")
    target_period_col = get_column_index(target_ws, "上期时间：XX年XX月XX日-XX年XX月XX日")
    if not all([target_code_col, target_period_col]):
        return 0
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    filled = 0
    for row in range(2, target_ws.max_row + 1):
        code = target_ws.cell(row=row, column=target_code_col).value
        if code:
            code_str = str(code).strip()
            if code_str in period_mapping:
                target_ws.cell(row=row, column=target_period_col, value=period_mapping[code_str]).border = thin_border
                target_ws.cell(row=row, column=target_period_col).alignment = Alignment(horizontal="center", vertical="center")
                filled += 1
    return filled


def copy_data_to_allocation_table(source_ws, target_ws):
    source_cols = {
        "报账点编码": get_column_index(source_ws, "报账点编码"),
        "报账点名称": get_column_index(source_ws, "报账点名称"),
        "电表编码": get_column_index(source_ws, "电表编码"),
        "上次分摊比例": get_column_index(source_ws, "上次分摊比例"),
        "实际分摊比例": get_column_index(source_ws, "实际分摊比例"),
        "缴费期始": get_column_index(source_ws, "缴费期始"),
    }
    target_cols = {
        "报账点编码": get_column_index(target_ws, "报账点编码"),
        "报账点名称": get_column_index(target_ws, "报账点名称"),
        "电表编码": get_column_index(target_ws, "电表编码"),
        "上期分摊比": get_column_index(target_ws, "上期分摊比"),
        "本期分摊比": get_column_index(target_ws, "本期分摊比"),
        "本期分摊开始时间：XX年XX月XX日": get_column_index(target_ws, "本期分摊开始时间：XX年XX月XX日"),
    }
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    max_row = source_ws.max_row
    for row in range(2, max_row + 1):
        target_row = row
        for key in ["报账点编码", "报账点名称", "电表编码"]:
            if source_cols[key]:
                val = source_ws.cell(row=row, column=source_cols[key]).value
                cell = target_ws.cell(row=target_row, column=target_cols[key], value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for key, target_key in [("上次分摊比例", "上期分摊比"), ("实际分摊比例", "本期分摊比")]:
            if source_cols[key]:
                val = source_ws.cell(row=row, column=source_cols[key]).value
                cell = target_ws.cell(row=target_row, column=target_cols[target_key], value=to_percent(val))
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
        if source_cols["缴费期始"] and target_cols["本期分摊开始时间：XX年XX月XX日"]:
            val = source_ws.cell(row=row, column=source_cols["缴费期始"]).value
            cell = target_ws.cell(row=target_row, column=target_cols["本期分摊开始时间：XX年XX月XX日"], value=format_date(val))
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
    return max_row - 1


def process_files(file1_path, file2_path, file3_path, file4_path, task_id=None):
    logs = []
    
    if task_id:
        task_output_folder = os.path.join(OUTPUT_FOLDER, task_id)
        os.makedirs(task_output_folder, exist_ok=True)
    else:
        task_output_folder = OUTPUT_FOLDER
    
    wb1 = openpyxl.load_workbook(file1_path)
    ws1 = wb1.active
    wb2 = openpyxl.load_workbook(file2_path)
    ws2 = wb2.active
    
    logs.append("文件读取成功")
    
    col_name = "缴费单编码"
    col1_idx = get_column_index(ws1, col_name)
    col2_idx = get_column_index(ws2, col_name)
    
    if col1_idx is None or col2_idx is None:
        logs.append("错误：未找到缴费单编码列")
        return {"success": False, "logs": logs}
    
    codes1 = get_payment_codes(ws1, col1_idx)
    codes2 = get_payment_codes(ws2, col2_idx)
    common_codes = codes1.intersection(codes2)
    
    logs.append(f"核实-分摊表: {len(codes1)} 个缴费单编码")
    logs.append(f"电表: {len(codes2)} 个缴费单编码")
    logs.append(f"共同存在: {len(common_codes)} 个")
    
    deleted1 = filter_rows_by_codes(ws1, col1_idx, common_codes)
    deleted2 = filter_rows_by_codes(ws2, col2_idx, common_codes)
    
    logs.append(f"从核实-分摊表删除 {deleted1} 行")
    logs.append(f"从电表删除 {deleted2} 行")
    
    output1 = os.path.join(task_output_folder, "核实-分摊表_过滤后.xlsx")
    output2 = os.path.join(task_output_folder, "电表_过滤后.xlsx")
    wb1.save(output1)
    wb2.save(output2)
    logs.append("过滤后文件已保存")
    
    alloc_wb, alloc_ws = create_allocation_table()
    data_rows = copy_data_to_allocation_table(ws2, alloc_ws)
    logs.append(f"分摊比例表基础数据: {data_rows} 行")
    
    county_mapping, station_mapping = build_mappings(ws1)
    filled_rows = vlookup_county_and_fill_city(ws1, alloc_ws, county_mapping)
    logs.append(f"区县/地市填充: {filled_rows} 行")
    
    station_rows = vlookup_station_code(alloc_ws, station_mapping)
    logs.append(f"站址编码填充: {station_rows} 行")
    
    if file3_path and os.path.exists(file3_path):
        wb3 = openpyxl.load_workbook(file3_path)
        station_name_mapping = build_station_name_mapping(wb3)
        name_rows = vlookup_station_name(alloc_ws, station_name_mapping)
        logs.append(f"铁塔站点名称填充: {name_rows} 行")
    
    if file4_path and os.path.exists(file4_path):
        wb4 = openpyxl.load_workbook(file4_path)
        ws4 = wb4.active
        period_mapping = build_payment_period_mapping(ws4)
        period_rows = vlookup_payment_period(alloc_ws, period_mapping)
        logs.append(f"上期时间填充: {period_rows} 行")
    
    output_alloc = os.path.join(task_output_folder, "分摊比例表.xlsx")
    alloc_wb.save(output_alloc)
    logs.append("分摊比例表已生成")
    
    return {"success": True, "logs": logs}


@manage_app.route('/')
def index():
    template_path = os.path.join(BASE_DIR, 'templates', 'manage', 'index.html')
    with open(template_path, 'r', encoding='utf-8') as f:
        return f.read()


@manage_app.route('/upload', methods=['POST'])
def upload():
    task_id = str(uuid.uuid4())
    task_upload_folder = os.path.join(UPLOAD_FOLDER, task_id)
    os.makedirs(task_upload_folder, exist_ok=True)
    
    files = {}
    for key in ['file1', 'file2', 'file3', 'file4']:
        if key in request.files:
            file = request.files[key]
            if file.filename:
                safe_name = secure_filename(file.filename)
                filepath = os.path.join(task_upload_folder, safe_name)
                file.save(filepath)
                files[key] = filepath
    return jsonify({"success": True, "files": files, "task_id": task_id})


@manage_app.route('/run', methods=['POST'])
def run_process():
    data = request.get_json()
    file1 = data.get('file1', '')
    file2 = data.get('file2', '')
    file3 = data.get('file3', '')
    file4 = data.get('file4', '')
    task_id = data.get('task_id', '')
    
    if not file1 or not file2:
        return jsonify({"success": False, "message": "请上传核实-分摊表和电表文件"})
    
    result = process_files(file1, file2, file3, file4, task_id)
    result['task_id'] = task_id
    return jsonify(result)


@manage_app.route('/list_files', methods=['GET'])
def list_files():
    task_id = request.args.get('task_id', '')
    if task_id:
        task_output_folder = os.path.join(OUTPUT_FOLDER, task_id)
        if not os.path.exists(task_output_folder):
            return jsonify({"files": []})
        target_folder = task_output_folder
    else:
        target_folder = OUTPUT_FOLDER
    
    files = []
    for f in os.listdir(target_folder):
        if f.endswith('.xlsx'):
            filepath = os.path.join(target_folder, f)
            size = os.path.getsize(filepath)
            files.append({
                "name": f,
                "size": f"{size / 1024:.1f} KB"
            })
    return jsonify({"files": files})


@manage_app.route('/download/<task_id>/<filename>')
def download(task_id, filename):
    safe_filename = secure_filename(filename)
    if '..' in task_id or '/' in task_id:
        return jsonify({"success": False, "message": "无效的请求"}), 400
    
    task_output_folder = os.path.join(OUTPUT_FOLDER, task_id)
    if not os.path.exists(task_output_folder):
        return jsonify({"success": False, "message": "文件不存在"}), 404
    
    return send_from_directory(task_output_folder, safe_filename, as_attachment=True)